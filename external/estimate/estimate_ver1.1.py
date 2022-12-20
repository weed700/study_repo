import sys
import pandas as pd
import os

# 오브젝트
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QAxContainer import *
from PyQt5.QtGui import * 
from PyQt5.QtCore import QDate, Qt

# 엑셀 저장
import openpyxl as oxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

# pdf 저장
import xlwings as xw


# form_class2 = uic.loadUi("s_window.ui")[0]

# class Est_Manager(QMainWindow, form_class2):

#     def __init__(self):
#         super().__init__()
#         self.setUI()
#         self.setWindowTitle("견적 담당자")
#         self.setWindowIcon(QIcon("gluesys_logo.png"))

form_class = uic.loadUiType("main_windows.ui")[0]

class MyWindow(QMainWindow, form_class):
    model_name = ''; detail_model_name = ''
    disc=0
    est_manager_list ={'이사 박 근 식':['010-5669-3172','gspark@gluesys.com','5371'], 
                       '이사 허 남 중':['010-3741-0869','njheo@gluesys.com','8158'],
                       '이사 김 재 호':['010-7520-9425','justin@gluesys.com','5302'],
                       '대리 우 동 현':['010-5769-7610','dhwoo@gluesys.com','5301'],
                       '대리 김 정':['010-3153-3327','jkim@gluesys.com','5374']}
    m_name=''; m_code=''; m_disc=''; m_count=''
    est_code_list = ['(N)', '(E)', '(P)', '(G)', '(A)', '(M)']

    def __init__(self):
        super().__init__()
        self.setUI()
        self.setWindowTitle("견적양식")
        self.setWindowIcon(QIcon("gluesys_logo.png"))
        #self.setFixedSize(1024, 600)

    # UI(window)에 프린트되는 아이템들
    def setUI(self):
        self.setupUi(self)  # window 창 
        self.Model_Check()  # 모델 체크 함수
        self.Manager()      # 견적 담당자
        self.Code()         # 견적 코드

        # 오늘 날짜 입력
        self.now = QDate.currentDate()
        self.date.setText(self.now.toString('yyyy년 MM월 dd일'))

        self.count_init.clicked.connect(self.Count_E_Init) # 수량 초기화 버튼
        self.OK.clicked.connect(self.Save_E_button) # 최종 버튼

    # 수량 초기화 버튼시 발생하는 이벤트 함수
    def Count_E_Init(self):
        #print("수량 초기화")
        for key in self.db_val.keys():
            self.money[key][1]=1
            self.db_val[key][1][0].setText('')
        self.Money_func()

    # 견적 담당자 이벤트 함수
    def Manager_E_func(self):
        a = self.sender()
        #print(a.currentText())
        self.m_name=a.currentText()
        self.Num_make()

    # 견적 담당자 함수
    def Manager(self):
        for m in self.est_manager_list.keys():
            self.est_manager.addItem(m)
        self.est_manager.currentTextChanged.connect(self.Manager_E_func)

    # 견적 코드 이벤트 함수
    def Code_E_func(self):
        a= self.sender()
        self.m_code=a.currentText()
        self.Num_make()

    # 견적 코드 함수
    def Code(self):
        for m in self.est_code_list:
            self.est_code.addItem(m)
        self.est_code.currentTextChanged.connect(self.Code_E_func)

    # 모델 DB 변수 저장
    def Excel_DB(self):
        db = pd.read_excel('AS500+PD 견적 양식.xlsm', engine='openpyxl',index_col=[1, 2, 3, 5, 6])
        model_name = []
        for split in db.index:
            model_name.append(split[0])

        # 같은 모델 찾아서 딕셔너리 만들기
        dic={}
        for t in set(model_name):
            l=[]
            for a in db.index:
                if t == a[0]:
                    l.append([a[1],a[2],a[3],a[4]])
            dic[t]=l
        del dic['모델']
        return dic

    # 세부 모델 DB 변수 저장
    def Excel_DB_Detail(self):
        db = pd.read_excel('AS500+PD 견적 양식.xlsm', engine='openpyxl',index_col=[2, 3, 5, 6])
        model_name = []
        for split in db.index:
            model_name.append(split[0])

        # 같은 모델 찾아서 딕셔너리 만들기
        dic={}
        #for t in set(model_name):
        l=[]
        l2=[]

        for a in db.index:
            n=''
            if 'GW Single' == a[0]:
                n = a[0]
                l.append([a[1],a[2],a[3]])
            elif 'GW Cluster' == a[0]:
                n = a[0]
                l2.append([a[1],a[2],a[3]])

            if 'GW Single' == n:
                dic[n]=l
            elif 'GW Cluster' == n:
                dic[n]=l2
            
        return dic

    # 견적 횟수 이벤트 함수
    def Count_E_func(self):
        c = self.sender()
        self.m_count=c.text()
        self.Num_make()

    # 각 오브젝트들을 저장하는 함수
    def Ob_db(self,db_key):
        self.db_val = {}
        self.money = {}
        
        label_list=['model_','memory_','disk1_','disk2_','nic1_','nic2_','hba1_','hba2_','w1_','w2_','cpu_','op_']
        for l in label_list:
            box= l+'box'
            a=[]

            a.append(self.findChildren(QComboBox,box))
            
            for i in range(0,4):
                if 0 == i:
                    a.append(self.findChildren(QLineEdit,l+str(i)))
                else:
                    t = l+str(i)
                    a.append(self.findChildren(QLabel, t))

            self.db_val[box] = a
            self.db_val[box][0][0].currentTextChanged.connect(lambda: self.combobox_event(db_key))
            self.db_val[box][1][0].textChanged.connect(self.lineedit_func)
        
        self.disc_box.currentTextChanged.connect(self.Disc_func)
        self.est_count.textChanged.connect(self.Count_E_func)
        # 초기화
        for key in self.db_val.keys():
            self.money[key]=[0,1]

    # 모델 리스트 클리어 함수
    def Combobox_clear(self):
        for box in self.db_val.keys():
            self.db_val[box][0][0].clear()
    
    # 가격 보여주는 함수
    def Money_func(self):
        sum_1=0
        sum_2=0

        for key in self.money.keys():
            a=int(self.money[key][0]*self.money[key][1])
            b=int(self.money[key][0]*self.money[key][1]*self.disc)
            sum_1+=a
            sum_2+=b
            for temp in self.db_val.keys():
                if key == temp:
                    self.db_val[key][2][0].setText(format(int(a),',d'))
                    self.db_val[key][3][0].setText(format(int(b),',d'))

        self.money_sum_1.setText(format(sum_1,',d'))
        self.money_sum_2.setText(format(sum_2,',d'))
        temp = int(sum_2/10000)*10000
        self.tt.setText(format(temp,',d'))  
    
    # 견적 번호 생성 함수
    def Num_make(self):
        # 견적 번호 만들기
        e_num=''
        self.m_disc=self.disc_box.currentText().split('%')
        if self.m_name and self.m_code and self.m_disc:
            e_num='GLS-'+self.now.toString('yyyyMMdd')+'0'+self.m_disc[0]+self.est_count.text()+'-'+self.est_manager_list[self.m_name][2]+self.m_code
        
        self.est_num.setText(e_num)

    # 할인율 함수
    def Disc_func(self):
        d=self.sender()
        val=d.currentText().split('%')
        self.disc=(100-int(val[0]))/100
        self.Money_func()
        self.Num_make()

    # 모델 리스트 설정
    def Combobox_list(self,db_key):
        for temp in self.est_db.keys():
            for i in range(0,len(self.est_db[temp])):
                # 모델
                if db_key == temp:
                    self.model_box.addItem(self.est_db[db_key][i][2])
                elif '메모리' == temp:
                    self.memory_box.addItem(self.est_db[temp][i][2])
                elif '디스크' == temp:
                    self.disk1_box.addItem(self.est_db[temp][i][2])
                    self.disk2_box.addItem(self.est_db[temp][i][2])
                elif 'NW' == temp:
                    if 'NIC' == self.est_db[temp][i][0]:
                        self.nic1_box.addItem(self.est_db[temp][i][2])
                        self.nic2_box.addItem(self.est_db[temp][i][2])
                    if 'HBA' == self.est_db[temp][i][0] or 'IB' == self.est_db[temp][i][0]:
                        self.hba1_box.addItem(self.est_db[temp][i][2])
                        self.hba2_box.addItem(self.est_db[temp][i][2])
                elif '워런티' == temp:
                    self.w1_box.addItem(self.est_db[temp][i][2])
                    # 나눠야하나?
                    self.w2_box.addItem(self.est_db[temp][i][2])
                    if '옵션SW' == self.est_db[temp][i][0]:
                        self.op_box.addItem(self.est_db[temp][i][2])
    
        # 할인율 0~100까지 리스트 만들기
        for i in range(0,101):
            d = str(i)+'%'
            self.disc_box.addItem(d)

    # 세부 모델 리스트
    def Combobox_detail_list(self,db_key):
        for i in range(0,len(self.detail_model_db[db_key])):
            if 'GW Single' == db_key:
                self.model_box.addItem(self.detail_model_db[db_key][i][1])
            elif 'GW Cluster' == db_key:
                self.model_box.addItem(self.detail_model_db[db_key][i][1])
            # elif 'CPU Upgrade' == db_key:
            #     self.cpu_box.addItem(self.detail_model_db[db_key][i][1])

        #self.cpu_box.currentTextChanged.connect(lambda: self.combobox_event(db_key))
        
    # 수량 입력 창
    def lineedit_func(self):
        for key in self.db_val.keys():
            if '' == self.db_val[key][1][0].text():
                a=1
            else:
                a=self.db_val[key][1][0].text()
            self.money[key][1] = int(a)
            
        self.Money_func()

    # 콤보박스 시그널 시 발생하는 함수(slot)
    def combobox_event(self,db_key):
        cb = self.sender()
        ob_name= cb.objectName()

        for key in self.est_db.keys(): 
            for i in range(0,len(self.est_db[key])):
                if db_key == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        v=self.est_db[key][i][3]
                        self.money['model_box'][0] = v
                        self.model_1.setText(format(v,',d'))
                        self.model_3.setText(self.est_db[key][i][1])          
                elif '메모리' == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        v=self.est_db[key][i][3]
                        self.money['memory_box'][0] = v
                        self.memory_1.setText(format(v,',d'))
                        self.memory_3.setText(self.est_db[key][i][1]) 
                elif '디스크' == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        v=self.est_db[key][i][3]
                        self.money[ob_name][0] = v
                        if ob_name == 'disk1_box':
                            self.disk1_1.setText(format(v,',d'))
                            self.disk1_3.setText(self.est_db[key][i][1])
                        elif ob_name == 'disk2_box':
                            self.disk2_1.setText(format(v,',d'))
                            self.disk2_3.setText(self.est_db[key][i][1]) 
                elif 'NW' == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        v=self.est_db[key][i][3]
                        self.money[ob_name][0] = v
                        if 'NIC' == self.est_db[key][i][0]:   
                            if ob_name == 'nic1_box':          
                                self.nic1_1.setText(format(v,',d'))
                                self.nic1_3.setText(self.est_db[key][i][1])
                            elif ob_name == 'nic2_box':
                                self.nic2_1.setText(format(v,',d'))
                                self.nic2_3.setText(self.est_db[key][i][1])
                        if 'HBA' == self.est_db[key][i][0] or 'IB' == self.est_db[key][i][0]:
                            if ob_name == 'hba1_box':
                                self.hba1_1.setText(format(v,',d'))
                                self.hba1_3.setText(self.est_db[key][i][1])
                            elif ob_name == 'hba2_box':
                                self.hba2_1.setText(format(v,',d'))
                                self.hba2_3.setText(self.est_db[key][i][1])
                elif '워런티' == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        v=self.est_db[key][i][3]
                        self.money[ob_name][0] = v
                        if ob_name == 'w1_box':
                            self.w1_1.setText(format(v,',d'))
                            self.w1_3.setText(self.est_db[key][i][1])
                        elif ob_name == 'w2_box':
                            self.w2_1.setText(format(v,',d'))
                            self.w2_3.setText(self.est_db[key][i][1])
                        if '옵션SW' == self.est_db[key][i][0]:
                            if ob_name == 'op_box':
                                self.op_1.setText(format(v,',d'))
                                self.op_3.setText(self.est_db[key][i][1])
                elif 'CPU Upgrade' == key:
                    if cb.currentText() == self.est_db[key][i][2]:
                        self.money[ob_name][0] = v
                        v=self.est_db[key][i][3]
                        self.cpu_1.setText(format(v,',d'))
                        self.cpu_3.setText(self.est_db[key][i][1])
        self.Money_func()     

    # 모델 체크 이벤트 시그널 함수
    def Model_Check(self):
        self.est_db=self.Excel_DB() # excel DB 딕셔너리로 변경하여 저장
        # 모델 딕셔너리(디스크, 메모리 등 제거)
        self.model_db = self.est_db.copy()
        del self.model_db['메모리']
        del self.model_db['디스크']
        del self.model_db['워런티']
        del self.model_db['NW']

        self.detail_model_db = self.Excel_DB_Detail() # excel DB detail 모델 부분 딕셔너리로 변경하여 저장
        print(self.detail_model_db.keys())
        # 라디오 버튼 생성
        x=10; y=20
        
        for b_name in sorted(self.model_db.keys()): 
            rad=QRadioButton(b_name,self.groupBox)
            rad.move(x,y)
            rad.name=b_name
            rad.clicked.connect(self.Check_Event) # 체크 됬을 때 보내는 시그널
            y+=50

        # 세부 모델 라디오 버튼 생성
        d_x=10; d_y=20
        for d_name in sorted(self.detail_model_db.keys()): 
            rad2=QRadioButton(d_name,self.groupBox_2)
            rad2.move(d_x,d_y)
            rad2.name=d_name
            rad2.clicked.connect(self.Check_Event_Detail) # 체크 됬을 때 보내는 시그널
            d_y+=30

    # 모델 버튼 클릭시 발생 함수
    def Check_Event(self):
        radi=self.sender()
        if radi.isChecked():
            self.Ob_db(radi.name)
            self.Combobox_clear() # 콤보박스 전부 클리어 함수
            self.Combobox_list(radi.name) # 콤보박스 리스트 설정

    # 세부 모델 버튼 클릭시 발생 함수
    def Check_Event_Detail(self):
        radi=self.sender()
        if radi.isChecked():
            self.model_box.clear() # 콤보박스 모델 부분 클리어 함수
            self.Combobox_detail_list(radi.name) # 콤보박스 리스트 설정

    #엑셀 생성 함수   
    def Excel_create(self):
        wb = oxl.Workbook()
        ws = wb.active

        BORDER_THIN = 'thin'

        # 병합 셀 만들기
        cell = {}

        for l in range(1, 52):
            c = []
            if 1 == l or 2 == l:
                c.append('D'+str(l)+':I'+str(l))
            elif 4 == l:
                c.append('A'+str(l)+':I'+str(l+1))
            elif 7 == l or 8 == l or 9 == l or 10 == l or 11 == l or 12 == l or 13 == l:
                c.append('A'+str(l)+':B'+str(l))
                c.append('C'+str(l)+':D'+str(l))
                if 10 == l or 12 == l:
                    c.append('H'+str(l)+':I'+str(l+1))
                    c.append('G'+str(l)+':G'+str(l+1))
                else:
                    c.append('H'+str(l)+':I'+str(l))
            elif 15 == l or 16 == l:
                c.append('C'+str(l)+':E'+str(l))
            elif 15 == l or 16 == l:
                c.append('C'+str(l)+':E'+str(l))
            elif 18 == l or 18 == l:
                c.append('A'+str(l)+':E'+str(l))
                c.append('H'+str(l)+':I'+str(l))
            elif 19 == l or 21 == l or 22 == l or 23 == l or 24 == l or 25 == l or 26 == l or 27 == l or 28 == l or 29 == l\
                        or 30 == l or 31 == l or 32 == l or 33 == l or 34 == l or 35 == l or 36 == l or 37 == l or 38 == l or 39 == l\
                            or 40 == l or 41 == l or 42 == l:
                c.append('B'+str(l)+':C'+str(l))
                c.append('D'+str(l)+':E'+str(l))
            elif 20 == l:
                c.append('A'+str(l)+':I'+str(l))
            elif 43 == l or 44 == l or 45 == l:
                c.append('H'+str(l)+':I'+str(l))
            elif 46 == l or 47 == l or 48 == l or 49 == l or 50 == l or 51 == l:
                c.append('A'+str(l)+':I'+str(l))

            if c:
                cell[l] = c

        # 병합
        data = []
        for i in cell.keys():
            for c in cell[i]:
                ws.merge_cells(c)

        # 알파벳 생성
        alp=[]
        for a in range(65,91):
            alp.append(chr(a))

        # 값 넣기
        for x in range(1,52):
            s=9; bold_f=False; hori='center'
            for y in alp:
                ce=y+str(x)
                val=''
                if 1 == x and 'D' == y:
                    val = '(14055) 테스트 안양시 동안구 시민대로327번길 11-31 파낙스R&D센터 5층'
                elif 2 == x and 'D' == y:
                    val = 'TEL : 070-8787-5376     FAX : 031-388-3261     http://www.gluesys.com'
                elif 4 == x and 'A' == y:
                    val = '見 積 書'
                    s=22; bold_f=True
                elif 7 == x and 'A' == y:
                    val = '견적번호'
                elif 7 == x and 'C' == y:
                    val = '$견적번호'
                elif 7 == x and 'G' == y:
                    val = '담당자'
                elif 7 == x and 'H' == y:
                    val = '$담당자'
                elif 8 == x and 'A' == y:
                    val = '수신'
                elif 8 == x and 'C' == y:
                    val = '$수신'
                elif 8 == x and 'G' == y:
                    val = '전화번호'
                elif 8 == x and 'H' == y:
                    val = '$전화번호'
                elif 9 == x and 'A' == y:
                    val = '참조'
                elif 9 == x and 'C' == y:
                    val = '$참조'
                elif 9 == x and 'G' == y:
                    val = 'E-mail'
                elif 9 == x and 'H' == y:
                    val = '$E-mail'
                elif 10 == x and 'A' == y:
                    val = '연락처'
                elif 10 == x and 'C' == y:
                    val = '$연락처'
                elif 10 == x and 'G' == y:
                    val = '상호'
                elif 10 == x and 'H' == y:
                    val = '㈜ 글 루 시 스'
                elif 11 == x and 'A' == y:
                    val = '견적일'
                elif 11 == x and 'C' == y:
                    val = '$견적일'
                elif 12 == x and 'A' == y:
                    val = '유효기간'
                elif 12 == x and 'C' == y:
                    val = '$유효기간'
                elif 12 == x and 'G' == y:
                    val = '대표이사'
                elif 12 == x and 'H' == y:
                    val = '박  성  순  (직인생략)'
                elif 13 == x and 'A' == y:
                    val = '납품가능일'
                elif 13 == x and 'C' == y:
                    val = '납품 일정 협의'
                elif 15 == x and 'A' == y:
                    val = '견적가'
                elif 15 == x and 'B' == y:
                    val = ':'
                elif 15 == x and 'C' == y:
                    val = '$일금영원정(VAT포함)'
                elif 16 == x and 'A' == y:
                    val = '건명'
                elif 16 == x and 'B' == y:
                    val = ':'
                elif 16 == x and 'C' == y:
                    val = '$건명을입력하세요'
                elif 18 == x and 'A' == y:
                    val = '귀사의 무궁한 발전을 기원하오며, 아래와 같이 견적합니다.'
                    hori='left'
                elif 18 == x and 'H' == y:
                    val = '단위. 원(V.A.T 별도)'
                elif 19 == x and 'A' == y:
                    val = '번호'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'B' == y:
                    val = '모델'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'D' == y:
                    val = '상세내역'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'F' == y:
                    val = '수량'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'G' == y:
                    val = '소비자단가'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'H' == y:
                    val = '공급단가'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 19 == x and 'I' == y:
                    val = '공급금액'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='D6DCE4', end_color='D6DCE4')
                elif 20 == x and 'A' == y:
                    val = '$.0 Usable'
                    ws[ce].fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
                    hori='left'

                if '' != val:
                    ws[ce] = val; ws[ce].font = Font(size=s, bold=bold_f); ws[ce].alignment = Alignment(horizontal=hori,vertical='center')


        box = Border(left=Side(border_style=BORDER_THIN,
                                    color='123456'),
                            right=Side(border_style=BORDER_THIN,
                                        color='123456'),
                            top=Side(border_style=BORDER_THIN,
                                    color='123456'),
                            bottom=Side(border_style=BORDER_THIN,
                                        color='123456'))
        box_lr = Border(left=Side(border_style=BORDER_THIN,
                                color='123456'),
                        right=Side(border_style=BORDER_THIN,
                                    color='123456'))
        box_top = Border(top=Side(border_style=BORDER_THIN,
                                color='123456'))
        box_bottom = Border(bottom=Side(border_style=BORDER_THIN,
                                color='123456'))

        # 양식 만들기
        for i in range(1,52):
            if 7 <= i and i <= 13:
                for a in 'A' 'B' 'C' 'D':
                    ws[a+str(i)].border = box
                for a in 'G' 'H' 'I':
                    ws[a+str(i)].border = box
            elif 19 <= i and i <=20:
                for a in 'A' 'B' 'C' 'D' 'E' 'F' 'G' 'H' 'I':
                    ws[a+str(i)].border = box
            elif 21 <= i and i <=42:
                for a in 'A' 'B' 'C' 'D' 'E' 'F' 'G' 'H' 'I':
                    ws[a+str(i)].border = box_lr
            elif 43 == i:
                for a in 'A' 'B' 'C' 'D' 'E' 'F':
                    ws[a+str(i)].border = box_top
                for a in 'G' 'H' 'I':
                    ws[a+str(i)].border = box
            elif 44 == i:
                for a in 'G' 'H' 'I':
                    ws[a+str(i)].border = box
            elif 45 == i:
                for a in 'G' 'H' 'I':
                    ws[a+str(i)].border = box
            elif 46 <= i and i <= 51:
                a = 'I'
                ws[a+str(i)].border = Border(right=Side(border_style=BORDER_THIN,
                                                color='123456'))
                if 46 == i:
                    for a in 'A' 'B' 'C' 'D' 'E' 'F' 'G' 'H':
                        ws[a+str(i)].border = box_top
                elif 51 == i:
                    for a in 'A' 'B' 'C' 'D' 'E' 'F' 'G' 'H' 'I':
                        if 'I' == a:
                            ws[a+str(i)].border = Border(right=Side(border_style=BORDER_THIN,
                                                color='123456'), 
                                                bottom=Side(border_style=BORDER_THIN,
                                                color='123456'))
                        else:
                            ws[a+str(i)].border = box_bottom

        # 이미지 넣기
        gluesys_logo = Image('gluesys.png')
        gluesys_logo.height = 37
        gluesys_logo.width = 126
        ws.add_image(gluesys_logo, 'A1')

        good_logo = Image('good.png')
        good_logo.height = 38
        good_logo.width = 82
        ws.add_image(good_logo, 'B28')

        # 셀 사이즈 조작
        ws.column_dimensions['A'].width = 6.5
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 11.75
        ws.column_dimensions['E'].width = 16.5
        ws.column_dimensions['F'].width = 3
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9.5

        # 셀 초기 선 삭제
        ws.sheet_view.showGridLines = False

        # sheet 이름
        ws.title = "estimate"
        save_filename="est_test.xlsx"
        
        # 엑셀 저장
        wb.save(save_filename)
        # PDF 저장
        self.PDF_create(save_filename)
    
    # PDF 생성 함수
    def PDF_create(self,filename):
        ### PDF 생성 함수
        # 파일 불러오기
        book = xw.Book(filename)

        # pdf 로 저장하기
        current_work_dir = os.getcwd()   # 현재 작업중인 폴더에 저장하기
        # 절대경로로 파일 위치 입력
        pdf_path = os.path.join(current_work_dir, "est_pdf_test.pdf") 

        # PDF 로 저장할 시트 선택하기(본 예제에서는 첫 번째 시트 선택하기)
        report_sheet = book.sheets[0]
        # PDF 로 저장하기
        report_sheet.api.ExportAsFixedFormat(0, pdf_path)

    # 최종 저장하는 버튼 이베튼 함수
    def Save_E_button(self):
        # 버튼 눌렀을 때 이벤트 결과 부분
        QMessageBox.information(self,'생성 확인 메시지','엑셀 파일 & PDF 생성')
        self.Excel_create()

    
if __name__ == "__main__":
    print("견적양식 프로그램 실행")
    app=QApplication(sys.argv)
    myApp = MyWindow()
    myApp.show()
    app.exec_()
